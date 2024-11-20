import pandas as pd
import openai
from datetime import datetime
import os
from openpyxl.styles import PatternFill

# Initialize the OpenAI client
client = openai.OpenAI(api_key=os.environ["OPENAI_API_KEY"])

def read_excel_files():
    """Read the input Excel files."""
    try:
        prompts = pd.read_excel('C:/Users/George/Dropbox/R_files/compost-tp/prompts.xlsx')
        value_dims = pd.read_excel('C:/Users/George/Dropbox/R_files/compost-tp/value_dims_compass.xlsx')
        scale_defs = pd.read_excel('C:/Users/George/Dropbox/R_files/compost-tp/multi-scale_definitions_compass.xlsx')
        
        # Create dimension-specific scale descriptions
        dimension_scales = {}
        for dimension in value_dims['value dimensions']:
            if dimension in scale_defs.columns:
                dimension_scales[dimension] = dict(zip(scale_defs['scale_point'], scale_defs[dimension]))
        
        # Convert value dimensions and their definitions to dictionary
        dim_definitions = dict(zip(value_dims['value dimensions'], value_dims['dim_definitions']))
        
        return prompts, dim_definitions, dimension_scales
    except Exception as e:
        print(f"Error reading Excel files: {e}")
        return None, None, None

def create_evaluation_prompt(text, dimension, dim_definition, dimension_scales):
    """Create a prompt for the OpenAI API."""
    # Get dimension-specific scale descriptions
    scale_descriptions = dimension_scales[dimension]
    
    # Create the scale description text
    scale_text = "\n".join([f"{point} = {desc}" for point, desc in sorted(scale_descriptions.items())])
    
    return f"""You will evaluate the following statement based on the dimension of {dimension}.

Dimension Definition:
{dim_definition}

Please rate the statement using a {len(scale_descriptions)}-point scale where:
{scale_text}

Statement: {text}

Please respond with ONLY a single number ({min(scale_descriptions.keys())}-{max(scale_descriptions.keys())}) representing your rating for {dimension}."""

def get_openai_rating(prompt, dimension, dimension_scales):
    """Get rating from OpenAI API."""
    scale_descriptions = dimension_scales[dimension]
    try:
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {
                    "role": "system", 
                    "content": f"You are a precise evaluator. Respond only with a single number between {min(scale_descriptions.keys())} and {max(scale_descriptions.keys())}."
                },
                {"role": "user", "content": prompt}
            ],
            temperature=0.3
        )
        rating = int(response.choices[0].message.content.strip())
        valid_points = scale_descriptions.keys()
        return rating if rating in valid_points else min(valid_points) + len(valid_points) // 2
    except Exception as e:
        print(f"Error getting OpenAI rating: {e}")
        return min(scale_descriptions.keys()) + len(scale_descriptions) // 2

def apply_color_formatting(writer, sheet_name, dimension_columns, dimension_scales):
    """Apply color formatting to the ratings."""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Create color schemes for each dimension
    for col, dimension in zip(dimension_columns, dimension_scales.keys()):
        scale_descriptions = dimension_scales[dimension]
        scale_points = sorted(list(scale_descriptions.keys()))
        num_points = len(scale_points)
        
        # Create color gradient from red to yellow to green
        colors = {}
        for i, point in enumerate(scale_points):
            if i < num_points / 2:
                # Red to yellow gradient for lower half
                red = 255
                green = int((i * 2 / num_points) * 255)
            else:
                # Yellow to green gradient for upper half
                red = int((2 - (i * 2 / num_points)) * 255)
                green = 255
            
            colors[point] = f'{red:02x}{green:02x}80'
        
        # Apply formatting for this dimension
        col_letter = chr(65 + col)
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet[f"{col_letter}{row}"]
            if cell.value in colors:
                cell.fill = PatternFill(start_color=colors[cell.value],
                                      end_color=colors[cell.value],
                                      fill_type='solid')

def main():
    # Read Excel files
    prompts, dim_definitions, dimension_scales = read_excel_files()
    if prompts is None or dim_definitions is None or dimension_scales is None:
        return
    
    # Get value dimensions
    dimensions = list(dim_definitions.keys())
    
    # Create timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Process each prompt and dimension
    for dimension in dimensions:
        ratings = []
        for prompt in prompts['Prompt']:
            evaluation_prompt = create_evaluation_prompt(
                prompt, 
                dimension, 
                dim_definitions[dimension],
                dimension_scales
            )
            rating = get_openai_rating(evaluation_prompt, dimension, dimension_scales)
            ratings.append(rating)
        
        # Add ratings column to prompts.xlsx
        column_name = f"{dimension}_Rating"
        prompts[column_name] = ratings
    
    # Add timestamp column
    prompts['Evaluation_Timestamp'] = timestamp
    
    # Calculate summary statistics
    summary = prompts[[col for col in prompts.columns if col.endswith('_Rating')]].agg(['mean', 'std', 'min', 'max'])
    
    # Create output file
    output_path = 'C:/Users/George/Dropbox/R_files/compost-tp/tp-results.xlsx'
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        prompts.to_excel(writer, sheet_name='Evaluations', index=False)
        summary.to_excel(writer, sheet_name='Summary')
        
        # Get the column indices and corresponding dimensions for ratings
        dimension_columns = [(list(prompts.columns).index(col), col.replace('_Rating', '')) 
                           for col in prompts.columns if col.endswith('_Rating')]
        
        # Apply color formatting
        apply_color_formatting(writer, 'Evaluations', 
                             [col for col, _ in dimension_columns], 
                             dimension_scales)

if __name__ == "__main__":
    main()
