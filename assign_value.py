import pandas as pd
import openai
from datetime import datetime
import os
from openpyxl.styles import PatternFill
from tkinter import filedialog
import tkinter as tk
from llm_handler import LLMHandler

def initialize_client():
    """Initialize LLMHandler with API key from .env file"""
    try:
        config = {
            'openai_api_key': None,
            'model': 'gpt-3.5-turbo',
            'temperature': 0.7,
            'max_tokens': 2000
        }
        
        # Load API key from .env file
        if os.path.exists('.env'):
            with open('.env', 'r') as f:
                for line in f:
                    if line.startswith('OPENAI_API_KEY='):
                        api_key = line.split('=')[1].strip()
                        config['openai_api_key'] = api_key
                        print(f"API key loaded from .env file: {api_key[:6]}...{api_key[-4:]}")
                        break
                else:
                    print("Error: OPENAI_API_KEY not found in .env file")
                    return None, None
        else:
            print("Error: .env file not found")
            return None, None

        # Initialize LLMHandler
        handler = LLMHandler(config)
        return handler, config
    except Exception as e:
        print(f"Error initializing LLMHandler: {str(e)}")
        return None, None

def validate_file_structure(df, file_type):
    """Validate the structure of uploaded files."""
    if file_type == "design_process":
        required_cols = ["Prompt"]
        if not all(col in df.columns for col in required_cols):
            return False, "Design Process Output file must contain a 'Prompt' column"
    
    elif file_type == "value_dimensions":
        required_cols = ["value dimensions", "dim_definitions"]
        if not all(col in df.columns for col in required_cols):
            return False, "Value Dimensions file must contain 'value dimensions' and 'dim_definitions' columns"
    
    elif file_type == "scale_definitions":
        if "scale_point" not in df.columns:
            return False, "Scale Definitions file must contain a 'scale_point' column"
        
    return True, "Valid file structure"

def read_excel_files():
    """Read the input Excel files using file dialog."""
    try:
        # Create root window and hide it
        root = tk.Tk()
        root.withdraw()
        
        # Dictionary to store file paths and their descriptions
        file_requests = [
            ("Design Process Output", "design_process"),
            ("Value Dimensions", "value_dimensions"),
            ("Scale Definitions for Value Dimensions", "scale_definitions")
        ]
        
        files_data = {}
        
        # Request each file
        for desc, file_type in file_requests:
            while True:
                filepath = filedialog.askopenfilename(
                    title=f"Select {desc} file (.xlsx)",
                    filetypes=[("Excel files", "*.xlsx")]
                )
                
                if not filepath:  # User cancelled
                    print("File selection cancelled. Exiting program.")
                    return None, None, None
                
                if not filepath.lower().endswith('.xlsx'):
                    print(f"Error: Please select a valid .xlsx file for {desc}")
                    continue
                
                try:
                    df = pd.read_excel(filepath)
                    is_valid, message = validate_file_structure(df, file_type)
                    
                    if not is_valid:
                        print(f"Error: {message}")
                        continue
                    
                    files_data[file_type] = df
                    break
                    
                except Exception as e:
                    print(f"Error reading {desc} file: {e}")
                    continue
        
        # Process the files as before
        prompts = files_data["design_process"]
        value_dims = files_data["value_dimensions"]
        scale_defs = files_data["scale_definitions"]
        
        # Create dimension-specific scale descriptions
        dimension_scales = {}
        for dimension in value_dims['value dimensions']:
            if dimension in scale_defs.columns:
                dimension_scales[dimension] = dict(zip(scale_defs['scale_point'], scale_defs[dimension]))
            else:
                print(f"Warning: Scale definitions missing for dimension '{dimension}'")
        
        # Convert value dimensions and their definitions to dictionary
        dim_definitions = dict(zip(value_dims['value dimensions'], value_dims['dim_definitions']))
        
        return prompts, dim_definitions, dimension_scales
        
    except Exception as e:
        print(f"Error processing files: {e}")
        return None, None, None

def create_evaluation_prompt(text, dimension, dim_definition, dimension_scales):
    """Create a prompt for the OpenAI API."""
    # Get dimension-specific scale descriptions
    scale_descriptions = dimension_scales[dimension]
    
    # Create the scale description text
    scale_text = "\n".join([f"{point} = {desc}" for point, desc in sorted(scale_descriptions.items())])
    
    return f"""You will evaluate the following statement based on the dimension of {dimension}.

Dimension Definition:
{dim_definition}

Please rate the statement using a {len(scale_descriptions)}-point scale where:
{scale_text}

Statement: {text}

Please respond with ONLY a single number ({min(scale_descriptions.keys())}-{max(scale_descriptions.keys())}) representing your rating for {dimension}."""

def get_openai_rating(handler, prompt, dimension, dimension_scales):
    """Get rating from OpenAI API using LLMHandler"""
    if handler is None:
        print("Debug: LLMHandler is not initialized.")
        return None

    # Add debug message to verify API key in use
    print(f"Using API key: {handler.config['openai_api_key'][:6]}...{handler.config['openai_api_key'][-4:]}")

    scale_descriptions = dimension_scales[dimension]
    try:
        messages = [
            {
                "role": "system", 
                "content": f"You are a precise evaluator. Respond only with a single number between {min(scale_descriptions.keys())} and {max(scale_descriptions.keys())}."
            },
            {"role": "user", "content": prompt}
        ]
        
        # Use the sync_client from LLMHandler
        response = handler.sync_client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=messages,
            temperature=0.7
        )
        
        rating = int(response.choices[0].message.content.strip())
        valid_points = scale_descriptions.keys()
        return rating if rating in valid_points else min(valid_points) + len(valid_points) // 2
    except Exception as e:
        print(f"Error getting OpenAI rating: {e}")
        return min(scale_descriptions.keys()) + len(scale_descriptions) // 2

def apply_color_formatting(writer, sheet_name, dimension_columns, dimension_scales):
    """Apply color formatting to the ratings."""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Create color schemes for each dimension
    for col, dimension in zip(dimension_columns, dimension_scales.keys()):
        scale_descriptions = dimension_scales[dimension]
        scale_points = sorted(list(scale_descriptions.keys()))
        num_points = len(scale_points)
        
        # Create color gradient from red to yellow to green
        colors = {}
        for i, point in enumerate(scale_points):
            if i < num_points / 2:
                # Red to yellow gradient for lower half
                red = 255
                green = int((i * 2 / num_points) * 255)
            else:
                # Yellow to green gradient for upper half
                red = int((2 - (i * 2 / num_points)) * 255)
                green = 255
            
            colors[point] = f'{red:02x}{green:02x}80'
        
        # Apply formatting for this dimension
        col_letter = chr(65 + col)
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet[f"{col_letter}{row}"]
            if cell.value in colors:
                cell.fill = PatternFill(start_color=colors[cell.value],
                                      end_color=colors[cell.value],
                                      fill_type='solid')

def main():
    # Initialize LLMHandler
    handler, config = initialize_client()
    if handler is None:
        print("Failed to initialize LLMHandler")
        return
        
    # Read Excel files
    prompts, dim_definitions, dimension_scales = read_excel_files()
    if prompts is None or dim_definitions is None or dimension_scales is None:
        return
    
    # Get value dimensions
    dimensions = list(dim_definitions.keys())
    
    # Create timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Process each prompt and dimension
    for dimension in dimensions:
        ratings = []
        for prompt in prompts['Prompt']:
            evaluation_prompt = create_evaluation_prompt(
                prompt, 
                dimension, 
                dim_definitions[dimension],
                dimension_scales
            )
            rating = get_openai_rating(handler, evaluation_prompt, dimension, dimension_scales)
            ratings.append(rating)
        
        # Add ratings column to prompts.xlsx
        column_name = f"{dimension}_Rating"
        prompts[column_name] = ratings
    
    # Add timestamp column
    prompts['Evaluation_Timestamp'] = timestamp
    
    # Calculate summary statistics
    summary = prompts[[col for col in prompts.columns if col.endswith('_Rating')]].agg(['mean', 'std', 'min', 'max'])
    
    # Create timestamp for filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_filename = f"design-process-output-scores_{timestamp}.xlsx"
    
    # Create root window and hide it (if not already created)
    root = tk.Tk()
    root.withdraw()
    
    # Prompt user for save location
    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialfile=default_filename,
        title="Save Results As"
    )
    
    if not output_path:  # User cancelled
        print("Save operation cancelled. Exiting program.")
        return
    
    # Save the file
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            prompts.to_excel(writer, sheet_name='Evaluations', index=False)
            summary.to_excel(writer, sheet_name='Summary')
            
            # Get the column indices and corresponding dimensions for ratings
            dimension_columns = [(list(prompts.columns).index(col), col.replace('_Rating', '')) 
                               for col in prompts.columns if col.endswith('_Rating')]
            
            # Apply color formatting
            apply_color_formatting(writer, 'Evaluations', 
                                 [col for col, _ in dimension_columns], 
                                 dimension_scales)
        print(f"Results saved successfully to: {output_path}")
        
    except Exception as e:
        print(f"Error saving results: {e}")

if __name__ == "__main__":
    main()
